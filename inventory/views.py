from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Sum, Count, F, ExpressionWrapper, DecimalField, Q
from django.db.models.functions import TruncMonth
from django.utils import timezone
from datetime import timedelta, date
from .models import (
    Category, Supplier, Product, StockMovement,
    PurchaseOrder, Notification, ProductVariant,
)
from .forms import (
    ProductForm, CategoryForm, SupplierForm, StockMovementForm,
    StockMovementQuickForm, PurchaseOrderForm, ProductVariantForm,
)
from django.http import HttpResponse
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from accounts.decorators import manager_or_admin_required, admin_required
from django.contrib.auth.decorators import login_required
import json


# ──────────────────────────────────────────────
# DASHBOARD
# ──────────────────────────────────────────────

@login_required
@manager_or_admin_required
def dashboard(request):
    from django.utils import timezone as tz
    today = tz.localdate()
    thirty_days_ago = today - timedelta(days=30)

    products_qs = Product.objects.filter(status="active")
    categories = Category.objects.all()

    all_products = list(products_qs)

    low_stock_items = [p for p in all_products if p.is_low_stock]
    overstock_items = [p for p in all_products if p.is_overstock]
    ok_stock_items  = [p for p in all_products if p.stock_status == "ok"]

    total_units  = sum(p.quantity for p in all_products)
    total_value  = sum(p.total_value for p in all_products)
    total_cost   = sum(p.cost_price * p.quantity for p in all_products)

    pending_orders = PurchaseOrder.objects.filter(status__in=["draft", "sent"]).count()
    recent_movements = StockMovement.objects.order_by("-date")[:15]
    unread_count = Notification.objects.filter(is_read=False).count()

    category_labels = []
    category_values = []
    for cat in categories:
        cp = [p for p in all_products if p.category_id == cat.pk]
        cv = sum(p.total_value for p in cp)
        if cv > 0:
            category_labels.append(cat.name)
            category_values.append(round(float(cv), 2))

    cats_json = json.dumps(category_labels)
    vals_json = json.dumps(category_values)

    context = {
        "products": products_qs,
        "all_products": all_products,
        "total_products": products_qs.count(),
        "active_products": products_qs.count(),
        "total_units": total_units,
        "unit_label": products_qs.first().unit if products_qs.first() else "units",
        "total_value": total_value,
        "total_cost": total_cost,
        "low_stock": len(low_stock_items),
        "overstock": len(overstock_items),
        "ok_stock": len(ok_stock_items),
        "low_stock_items": low_stock_items[:10],
        "overstock_items": overstock_items,
        "pending_orders": pending_orders,
        "recent_movements": recent_movements,
        "unread_count": unread_count,
        "unread_nots": unread_count,
        "category_labels": cats_json,
        "category_values": vals_json,
        "low_threshold": 10,
        "overstock_threshold": 1000,
    }
    return render(request, "inventory/dashboard.html", context)


# ──────────────────────────────────────────────
# PRODUCTS
# ──────────────────────────────────────────────

@login_required
@manager_or_admin_required
def product_list(request):
    q        = request.GET.get("q", "")
    cat_id   = request.GET.get("category", "")
    status_f = request.GET.get("status", "")

    qs = Product.objects.all()

    if q:
        qs = qs.filter(
            Q(name__icontains=q) | Q(sku__icontains=q) | Q(barcode__icontains=q)
        )
    if cat_id:
        qs = qs.filter(category_id=cat_id)
    if status_f:
        qs = qs.filter(status=status_f)

    from django.core.paginator import Paginator
    paginator = Paginator(qs, 12)
    products  = paginator.get_page(request.GET.get("page"))

    return render(request, "inventory/product_list.html", {
        "products": products,
        "categories": Category.objects.all(),
        "category_id": cat_id,
        "query": q,
    })


@login_required
@manager_or_admin_required
def product_create(request):
    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES)
        if form.is_valid():
            p = form.save()
            messages.success(request, f"Product '{p.name}' created successfully.")
            return redirect("product_list")
    else:
        form = ProductForm()
    return render(request, "inventory/product_form.html", {"form": form, "title": "Add Product"})


@login_required
@manager_or_admin_required
def product_edit(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        form = ProductForm(request.POST, request.FILES, instance=product)
        if form.is_valid():
            prod = form.save()
            messages.success(request, f"Product '{prod.name}' updated.")
            return redirect("product_detail", pk=prod.pk)
    else:
        form = ProductForm(instance=product)
    return render(request, "inventory/product_form.html",
                  {"form": form, "product": product, "title": f"Edit {product.name}"})


@login_required
@manager_or_admin_required
def product_detail(request, pk):
    product = get_object_or_404(Product.objects.prefetch_related("variants", "movements"), pk=pk)
    in_qty_total = product.movements.filter(movement_type="IN").aggregate(t=Sum("quantity"))["t"] or 0
    out_qty_total = product.movements.filter(movement_type="OUT").aggregate(t=Sum("quantity"))["t"] or 0
    movements = product.movements.all().order_by("-date")
    return render(request, "inventory/product_detail.html", {
        "product": product,
        "in_qty_total": in_qty_total,
        "out_qty_total": out_qty_total,
        "movements": movements,
    })


@login_required
@manager_or_admin_required
def product_delete(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        product.status = "archived"
        product.save()
        messages.success(request, f"'{product.name}' archived.")
        return redirect("product_list")
    return render(request, "inventory/product_confirm_delete.html", {"product": product})


@login_required
@manager_or_admin_required
def product_archive(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        product.status = "archived"
        product.save()
        messages.success(request, f"'{product.name}' has been archived.")
    return redirect("product_list")


# ──────────────────────────────────────────────
# VARIANTS
# ──────────────────────────────────────────────

@login_required
@manager_or_admin_required
def variant_add(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        form = ProductVariantForm(request.POST)
        if form.is_valid():
            var = form.save(commit=False)
            var.product = product
            var.save()
            messages.success(request, "Variant added!")
            return redirect("product_detail", pk=pk)
    else:
        form = ProductVariantForm()
    return render(request, "inventory/variant_form.html",
                  {"form": form, "product": product, "title": "Add Variant"})


# ──────────────────────────────────────────────
# STOCK MOVEMENTS
# ──────────────────────────────────────────────

@login_required
@manager_or_admin_required
def stock_add(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        form = StockMovementForm(request.POST)
        if form.is_valid():
            qty   = form.cleaned_data["quantity"]
            reason = form.cleaned_data["reason"]
            notes = form.cleaned_data.get("notes", "")
            product.quantity += qty
            product.save()
            StockMovement.objects.create(
                product=product, movement_type="IN", quantity=qty,
                reason=reason, notes=notes, recorded_by=request.user.username,
            )
            messages.success(request, f"+ {qty} {product.unit} added to '{product.name}'.")
            return redirect("product_detail", pk=pk)
    else:
        form = StockMovementForm(initial={"product": product, "movement_type": "IN"})
    return render(request, "inventory/stock_form.html",
                  {"form": form, "product": product, "action": "Add Stock"})


@login_required
@manager_or_admin_required
def stock_remove(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        form = StockMovementForm(request.POST)
        if form.is_valid():
            qty   = form.cleaned_data["quantity"]
            reason = form.cleaned_data["reason"]
            notes  = form.cleaned_data.get("notes", "")
            if qty > product.quantity:
                messages.error(request, f"Not enough stock! Only {product.quantity} available.")
            else:
                product.quantity -= qty
                product.save()
                StockMovement.objects.create(
                    product=product, movement_type="OUT", quantity=qty,
                    reason=reason, notes=notes, recorded_by=request.user.username,
                )
                messages.success(request, f"− {qty} {product.unit} removed from '{product.name}'.")
            return redirect("product_detail", pk=pk)
    else:
        form = StockMovementForm(initial={"product": product, "movement_type": "OUT"})
    return render(request, "inventory/stock_form.html",
                  {"form": form, "product": product, "action": "Remove Stock"})


@login_required
@manager_or_admin_required
def stock_adjust(request, pk):
    product = get_object_or_404(Product, pk=pk)
    if request.method == "POST":
        form = StockMovementForm(request.POST)
        if form.is_valid():
            qty   = form.cleaned_data["quantity"]
            reason = form.cleaned_data["reason"]
            notes  = form.cleaned_data.get("notes", "")
            product.quantity += qty
            product.save()
            StockMovement.objects.create(
                product=product, movement_type="ADJUST", quantity=qty,
                reason=reason, notes=notes, recorded_by=request.user.username,
            )
            messages.success(request, f"Stock adjusted by {qty} for '{product.name}'.")
            return redirect("product_detail", pk=pk)
    else:
        form = StockMovementForm(initial={"product": product, "movement_type": "ADJUST"})
    return render(request, "inventory/stock_form.html",
                  {"form": form, "product": product, "action": "Adjust Stock"})


@login_required
@manager_or_admin_required
def stock_movement_quick(request):
    if request.method == "POST":
        form = StockMovementQuickForm(request.POST)
        if form.is_valid():
            product        = form.cleaned_data["product"]
            movement_type  = form.cleaned_data["movement_type"]
            qty            = form.cleaned_data["quantity"]
            reason         = form.cleaned_data["reason"]
            notes          = form.cleaned_data.get("notes", "")

            if movement_type in ("OUT", "RETURN_OUT"):
                if qty > product.quantity:
                    messages.error(request, f"Only {product.quantity} available.")
                    return redirect("dashboard")
                product.quantity -= qty
            else:
                product.quantity += qty
            product.save()

            StockMovement.objects.create(
                product=product, movement_type=movement_type, quantity=qty,
                reason=reason, notes=notes, recorded_by=request.user.username,
            )
            messages.success(request, f"Movement recorded for '{product.name}'.")
    return redirect("dashboard")


@login_required
@manager_or_admin_required
def movement_list(request):
    q          = request.GET.get("q", "")
    mtype      = request.GET.get("type", "")
    product_id = request.GET.get("product", "")

    qs = StockMovement.objects.select_related("product").all()
    if q:
        qs = qs.filter(Q(product__name__icontains=q) | Q(reason__icontains=q))
    if mtype:
        qs = qs.filter(movement_type=mtype)
    if product_id:
        qs = qs.filter(product_id=product_id)

    from django.core.paginator import Paginator
    paginator = Paginator(qs, 25)
    page      = paginator.get_page(request.GET.get("page"))
    return render(request, "inventory/movement_list.html", {
        "movements": page, "products": Product.objects.filter(status="active"),
        "movement_types": StockMovement.MOVEMENT_TYPES,
        "q": q, "mtype": mtype,
    })


# ──────────────────────────────────────────────
# CATEGORIES
# ──────────────────────────────────────────────

@login_required
@manager_or_admin_required
def category_list(request):
    categories = Category.objects.all()
    form = CategoryForm()
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Category added!")
            return redirect("category_list")
    return render(request, "inventory/category_list.html",
                  {"categories": categories, "form": form})


@login_required
@manager_or_admin_required
def category_edit(request, pk):
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "Category updated.")
            return redirect("category_list")
    else:
        form = CategoryForm(instance=category)
    return render(request, "inventory/category_form.html",
                  {"form": form, "category": category, "title": "Edit Category"})


@login_required
@manager_or_admin_required
def category_create(request):
    return category_list(request)


@login_required
@manager_or_admin_required
def category_delete(request, pk):
    cat = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        cat.delete()
        messages.success(request, "Category deleted.")
    return redirect("category_list")


# ──────────────────────────────────────────────
# SUPPLIERS
# ──────────────────────────────────────────────

@login_required
@manager_or_admin_required
def supplier_list(request):
    qs = Supplier.objects.all()
    search = request.GET.get("q", "")
    if search:
        qs = qs.filter(Q(name__icontains=search) | Q(contact_person__icontains=search))
    from django.core.paginator import Paginator
    paginator = Paginator(qs, 15)
    suppliers = paginator.get_page(request.GET.get("page"))
    form = SupplierForm()
    if request.method == "POST":
        form = SupplierForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Supplier added!")
            return redirect("supplier_list")
    return render(request, "inventory/supplier_list.html", {
        "suppliers": suppliers, "form": form, "query": search,
    })


@login_required
@manager_or_admin_required
def supplier_edit(request, pk):
    supplier = get_object_or_404(Supplier, pk=pk)
    if request.method == "POST":
        form = SupplierForm(request.POST, instance=supplier)
        if form.is_valid():
            form.save()
            messages.success(request, "Supplier updated.")
            return redirect("supplier_list")
    else:
        form = SupplierForm(instance=supplier)
    return render(request, "inventory/supplier_form.html",
                  {"form": form, "supplier": supplier, "title": "Edit Supplier"})


@login_required
@manager_or_admin_required
def supplier_create(request):
    return supplier_list(request)


@login_required
@manager_or_admin_required
def supplier_delete(request, pk):
    sup = get_object_or_404(Supplier, pk=pk)
    if request.method == "POST":
        sup.delete()
        messages.success(request, "Supplier deleted.")
    return redirect("supplier_list")


# ──────────────────────────────────────────────
# PURCHASE ORDERS
# ──────────────────────────────────────────────

@login_required
@manager_or_admin_required
def purchase_order_list(request):
    qs = PurchaseOrder.objects.select_related("supplier", "product").all()
    st = request.GET.get("status", "")
    if st:
        qs = qs.filter(status=st)
    supplier_id = request.GET.get("supplier", "")
    if supplier_id:
        qs = qs.filter(supplier_id=supplier_id)
    from django.core.paginator import Paginator
    paginator = Paginator(qs, 15)
    orders = paginator.get_page(request.GET.get("page"))
    return render(request, "inventory/purchase_order_list.html", {
        "orders": orders, "status_filter": st, "supplier_id": supplier_id,
        "suppliers": Supplier.objects.filter(is_active=True),
        "STATUS_CHOICES": PurchaseOrder.STATUS_CHOICES,
    })


@login_required
@manager_or_admin_required
def purchase_order_create(request):
    if request.method == "POST":
        form = PurchaseOrderForm(request.POST)
        if form.is_valid():
            order = form.save(commit=False)
            order.total_amount = order.quantity * order.unit_price
            order.order_number = "PO-{}".format(
                timezone.now().strftime("%Y%m%d%H%M%S")
            )
            order.save()
            messages.success(request, f"Purchase Order '{order.order_number}' created.")
            return redirect("purchase_order_list")
    else:
        form = PurchaseOrderForm()
    return render(request, "inventory/purchase_order_form.html",
                  {"form": form, "title": "Create Purchase Order"})


@login_required
@manager_or_admin_required
def purchase_order_receive(request, pk):
    order = get_object_or_404(PurchaseOrder, pk=pk)
    if request.method == "POST":
        order.status   = "received"
        order.received_date = timezone.now()
        order.save()
        product = order.product
        product.quantity += order.quantity
        product.save()
        StockMovement.objects.create(
            product=product, movement_type="IN", quantity=order.quantity,
            reason="delivery", notes=f"PO: {order.order_number}",
            recorded_by=request.user.username,
        )
        Notification.objects.create(
            notification_type="order_received",
            title   = f"Order {order.order_number} received",
            message = f"{order.quantity} units of {product.name} added to stock.",
            product = product,
        )
        messages.success(request,
                         f"Order {order.order_number} received — {order.quantity} units added.")
        return redirect("purchase_order_list")
    return render(request, "inventory/purchase_order_confirm.html", {"order": order})


@login_required
@manager_or_admin_required
def purchase_order_edit(request, pk):
    order = get_object_or_404(PurchaseOrder, pk=pk)
    if request.method == "POST":
        form = PurchaseOrderForm(request.POST, instance=order)
        if form.is_valid():
            o = form.save(commit=False)
            o.total_amount = o.quantity * o.unit_price
            o.save()
            messages.success(request, f"Order '{o.order_number}' updated.")
            return redirect("purchase_order_list")
    else:
        form = PurchaseOrderForm(instance=order)
    return render(request, "inventory/purchase_order_form.html",
                  {"form": form, "order": order, "title": f"Edit {order.order_number}"})


@login_required
@manager_or_admin_required
def purchase_order_cancel(request, pk):
    order = get_object_or_404(PurchaseOrder, pk=pk)
    if request.method == "POST":
        order.status = "cancelled"
        order.save()
        messages.success(request, f"Order '{order.order_number}' cancelled.")
    return redirect("purchase_order_list")


# ──────────────────────────────────────────────
# REPORTS
# ──────────────────────────────────────────────

@login_required
@manager_or_admin_required
def reports_view(request):
    products      = Product.objects.filter(status="active")
    cash_value    = sum(p.total_value for p in products)
    total_margin  = sum((p.price - p.cost_price) * p.quantity for p in products)

    movements = StockMovement.objects.all()
    movements = movements.order_by("-date")[:200]

    total_in  = sum(m.quantity for m in StockMovement.objects.filter(movement_type="IN"))
    total_out = sum(m.quantity for m in StockMovement.objects.filter(movement_type="OUT"))

    top_movers      = StockMovement.objects.filter(movement_type="IN").values("product__name").annotate(
        sold=Sum("quantity")
    ).order_by("-sold")[:5]
    slow_movers     = products.filter(quantity__gt=0).order_by("quantity")[:5]

    monthly = list(
        PurchaseOrder.objects
        .annotate(month=TruncMonth("order_date"))
        .values("month")
        .annotate(total=Sum("total_amount"))
        .order_by("month")[:6]
    )

    return render(request, "inventory/reports.html", {
        "products": products,
        "cash_value": cash_value,
        "total_margin": total_margin,
        "movements": movements,
        "total_in": total_in,
        "total_out": total_out,
        "top_movers": top_movers,
        "slow_movers": slow_movers,
        "monthly_orders": monthly,
        "categories": Category.objects.all(),
    })


@login_required
@manager_or_admin_required
def export_excel(request):
    products = Product.objects.filter(status="active")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Stock Valuation"
    headers = ["#", "Product Name", "SKU", "Category", "Supplier",
               "Unit", "Quantity", "Cost Price", "Sale Price", "Total Value", "Stock Status"]
    ws.append(headers)
    hfill = PatternFill(start_color="4f46e5", end_color="4f46e5", fill_type="solid")
    hfont = Font(bold=True, color="FFFFFF")
    thin  = Side(style="thin", color="dddddd")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
    alt_fill = PatternFill(start_color="f8f9ff", end_color="f8f9ff", fill_type="solid")
    for i, p in enumerate(products, 1):
        ws.append([
            i, p.name, p.sku or "—", p.category.name if p.category else "—",
            p.supplier.name if p.supplier else "—",
            p.get_unit_display(), p.quantity,
            float(p.cost_price), float(p.price), float(p.total_value), p.stock_status.upper(),
        ])
        for cell in ws[i + 1]:
            cell.border = border
            if i % 2 == 0:
                cell.fill = alt_fill
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        col_letter = get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = max(10, min(max_len + 2, 50))

    resp = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = "attachment; filename=stock_valuation.xlsx"
    wb.save(resp)
    return resp


@login_required
@manager_or_admin_required
def export_pdf(request):
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import inch

    products = Product.objects.filter(status="active")
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=stock_valuation.pdf"
    doc = SimpleDocTemplate(response, pagesize=A4, topMargin=0.75 * inch)
    styles = getSampleStyleSheet()
    title = Paragraph("<b>StockPro — Stock Valuation Report</b>", styles["Title"])
    story = [title, Spacer(1, 0.2 * inch)]
    data = [["#", "Product", "SKU", "Category", "Qty", "Cost", "Price", "Total", "Status"]]
    for i, p in enumerate(products, 1):
        data.append([
            i, p.name, p.sku or "—",
            p.category.name if p.category else "—",
            p.quantity,
            f"${p.cost_price:.2f}",
            f"${p.price:.2f}",
            f"${p.total_value:.2f}",
            p.stock_status.upper(),
        ])
    tbl = Table(data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ("BACKGROUND",   (0, 0), (-1, 0), colors.HexColor("#4f46e5")),
        ("TEXTCOLOR",    (0, 0), (-1, 0), colors.white),
        ("FONTNAME",     (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",     (0, 0), (-1, 0), 9),
        ("ALIGN",        (0, 0), (-1, 0), "CENTER"),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f8f9ff")]),
        ("GRID",         (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE",     (0, 1), (-1, -1), 8),
        ("TOPPADDING",   (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING",(0, 0), (-1, -1), 4),
    ]))
    story.append(tbl)
    doc.build(story)
    return response


# ──────────────────────────────────────────────
# ALERTS / NOTIFICATIONS
# ──────────────────────────────────────────────

@login_required
@manager_or_admin_required
def alerts_list(request):
    notes = Notification.objects.order_by("-created_at")
    if request.method == "POST":
        pk = request.POST.get("pk", "")
        action = request.POST.get("action", "")
        n = get_object_or_404(Notification, pk=pk)
        if action == "mark_read":
            n.is_read = True; n.save()
            messages.success(request, "Notification marked as read.")
        elif action == "delete":
            n.delete()
            messages.success(request, "Notification deleted.")
        return redirect("alerts")
    return render(request, "inventory/alerts.html", {"notifications": notes})


@login_required
@manager_or_admin_required
def mark_notification_read(request, pk):
    note = get_object_or_404(Notification, pk=pk)
    note.is_read = True
    note.save()
    return redirect("alerts")
